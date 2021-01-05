# encoding: UTF-8
# coding: UTF-8

import requests
import openpyxl

# 注册请求
# url_reg = "http://120.78.128.25:8766/futureloan/member/register"                               # 请求行
# headers_reg = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}        # 请求头
# data_reg = {"mobile_phone":"15501030824","pwd":"lemontree0123456","type":"0","reg_name":"hi"}  # 请求正文
# result_reg = requests.post(url=url_reg,headers=headers_reg,json=data_reg)
# print(result_reg)                   # 返回http状态码
# print(result_reg.status_code)       # 返回http状态码
# print(result_reg.headers)           # 返回响应头
# print(result_reg.json())            # 获取响应正文
# print(result_reg.text)              # 获取响应正文

# 登录请求
# url_login = "http://120.78.128.25:8766/futureloan/member/login"
# headers_login = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
# data_login = {"mobile_phone": "15501030824","pwd":"lemontree0123456"}
# res_login = requests.post(url=url_login,headers=headers_login,json=data_login).json()
# print(res_login)
# url_recharge = "http://120.78.128.25:8766/futureloan/member/recharge"
# Token = res_login['data']['token_info']['token']
# recharge_id = res_login['data']['id']
# headers_recharge = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json","Authorization":"Bearer "+Token}
# data_recharge = {"member_id":recharge_id,"amount":500000}
# res_recharge = requests.post(url=url_recharge,headers=headers_recharge,json=data_recharge).text
# print(res_recharge)




# 自定义一个函数，形参包括位置参数／默认参数／不定长参数
# def good_job(salary, bonus, *args, subsidy=500, **kwargs):
#     sum1 = salary + bonus + subsidy
#     for i in args:
#         sum1 += i
#     for j in kwargs.values():
#         sum1 += j
#     return sum1  # 以上为函数的定义

# in_total = good_job(100000, 500000, 200000, 100000, others=5000)  # 此处为函数的调用
# print(in_total)
# print(good_job(100000, 500000, 200000, 100000, others=5000))



# 函数 -- openpyxl读取测试用例，并把每条测试用例以字典（键值对）类型保存在列表里
import openpyxl
def read_case(wb_name,sh_name):
    wb = openpyxl.load_workbook(wb_name)
    sh = wb[sh_name]
    rowmax = sh.max_row
    list_reg = []
    for i in range(2, rowmax+1) :
        dict_reg = dict(case_id = sh.cell(row=i,column=1).value,
        url_reg = sh.cell(row=i,column=5).value,
        data_reg = sh.cell(row=i,column=6).value,
        expect_reg = sh.cell(row=i,column=7).value)
        list_reg.append(dict_reg)
    return list_reg

# case_of_reg = read_case('test_case_api.xlsx','register')
# print(case_of_reg)



# 函数 -- requests发送注册或登录接口请求
import requests
def api_test(url1,data1):
    headers1 = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res_api_test = requests.post(url=url1,headers=headers1,json=data1).json()
    return res_api_test
#
# url_login = "http://120.78.128.25:8766/futureloan/member/login"
# data_login = {"mobile_phone": "15501030824","pwd":"lemontree0123456"}
# res_api_login = api_test(url1=url_login,data1=data_login)
# print(res_api_login)


# 函数 -- openpyxl写入断言结果
def write_res(wb_name,sh_name,row,column,final_res):
    wb = openpyxl.load_workbook(wb_name)
    sh = wb[sh_name]
    sh.cell(row=row,column=column).value = final_res
    wb.save(wb_name)


cases = read_case('/Users/tiancai/Downloads/lemon/Lemonclass-all projects/Project 3-API test/test_case01.xlsx','login')
for case in cases:
    case_id = case['case_id']
    url = case['url_reg']
    data = eval(case['data_reg'])
    expect = eval(case['expect_reg'])
    expect_code = expect['code']
    expect_msg = expect['msg']
    real_res_api_reg = api_test(url1=url,data1=data)
    real_code = real_res_api_reg['code']
    real_msg = real_res_api_reg['msg']
    print("期望结果为：{},{}".format(expect_code,expect_msg))
    print("实际结果为：{},{}".format(real_code,real_msg))
    if expect_code == real_code and expect_msg == real_msg:
        print('第{}条用例执行通过'.format(case_id))
        final_res = 'passed'
    else:
        print('第{}条用例执行不通过'.format(case_id))
        final_res ='failed'
    print('*'*20)
    write_res('/Users/tiancai/Downloads/lemon/Lemonclass-all projects/Project 3-API test/test_case01.xlsx','login', case_id+1, 8, final_res)